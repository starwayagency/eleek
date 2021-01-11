from django.utils.feedgenerator import (
  Atom1Feed,
  Rss201rev2Feed ,
  RssUserland091Feed ,
  Atom1Feed,
  SyndicationFeed,
)

from django.contrib.syndication.views import Feed
from django.utils.xmlutils import SimplerXMLGenerator
from .models import Item, ItemCategory, ItemFeature


from django.http import HttpResponse 

from openpyxl import Workbook
from datetime import datetime 


from box.core.sw_imp_exp.main import ExportMixin
from django.utils.html import strip_tags


def get_products_sheet():
  fieldnames = {
    'Код_товара':'',
    'Название_позиции':'',
    'Поисковые_запросы':'',
    'Описание':'',
    'Тип_товара':'',
    'Цена':'',
    'Цена от':'',
    'Ярлык':'',
    'HTML_заголовок':'',
    'HTML_описание':'',
    'HTML_ключевые_слова':'',
    'Валюта':'',
    'Скидка':'',
    'Cрок действия скидки от':'',
    'Cрок действия скидки до':'',
    'Единица_измерения':'',
    'Минимальный_объем_заказа':'',
    'Оптовая_цена':'',
    'Минимальный_заказ_опт':'',
    'Ссылка_изображения':'',
    'Наличие':'',
    'Количество':'',
    'Производитель':'',
    'Страна_производитель':'',
    'Номер_группы':'',
    'Адрес_подраздела':'',
    'Возможность_поставки':'',
    'Срок_поставки':'',
    'Способ_упаковки':'',
    'Личные_заметки':'',
    'Продукт_на_сайте':'',
    'Идентификатор_товара':'',
    'Уникальный_идентификатор':'',
    'Идентификатор_подраздела':'',
    'Идентификатор_группы':'',
    'Подарки':'',
    'Сопутствующие':'',
    'ID_группы_разновидностей':'',
    'Название_Характеристики':'',
    'Измерение_Характеристики':'',
    'Значение_Характеристики':'',
    'Название_Характеристики':'',
    'Измерение_Характеристики':'',
    'Значение_Характеристики':'',
  }
 
  return fieldnames

def get_categories_sheet():
  fieldnames = {
    'Номер_группы':'',
    'Название_группы':'',
    'Идентификатор_группы':'',
    'Номер_родителя':'',
    'Идентификатор_родителя':'',
    'HTML_заголовок_группы':'',
    'HTML_описание_группы':'',
    'HTML_ключевые_слова_группы':'',
  }
 
  return fieldnames

units = [
  '2г',
  '5г',
  '10 г',
  '50 г',
  '100г',
  '10см',
  'шт.',
  '10 шт.',
  '20 шт.',
  '50 шт.',
  '100 шт.',
  'ампула',
  'баллон',
  'банка',
  'блистер',
  'бобина',
  'бочка',
  'бут',
  'бухта',
  'ватт',
  'ведро',
  'выезд',
  'г',
  'га',
  'гигакалория',
  'год',
  'гр/кв.м',
  'дал',
  'два месяца',
  'день',
  'доза',
  'ед.',
  'кВт',
  'канистра',
  'карат',
  'кв.дм',
  'кв.м',
  'кв.см',
  'кв.фут',
  'квартал',
  'кг',
  'кг/кв.м',
  'км',
  'комплект',
  'коробка',
  'куб.дм',
  'куб.м',
  'л',
  'лист',
  'м',
  'мВт',
  'месяц',
  'мешок',
  'минута',
  'мл',
  'мм',
  'моток',
  'набор',
  'неделя',
  'номер',
  'объект',
  'паллетоместо',
  'пара',
  'партия',
  'пач',
  'пог.м',
  'полгода',
  'посевная единица',
  'птицеместо',
  'рейс',
  'рулон',
  'секция',
  'слово',
  'см',
  'смена',
  'сотка',
  'стакан',
  'страница',
  'сутки',
  'т',
  'т.у.шт.',
  'т/км',
  'таблетка',
  'тираж',
  'тыс. шт.',
  'тысяча',
  'тюбик',
  'упаковка',
  'услуга',
  'флакон',
  'час',
  'чел.',
  'шприц-туба',
  'ярд',
  'ящик',
]

def create_worksheet_with_items( workbook, items):
  worksheet1 = workbook.create_sheet(
      title='Export Products Sheet',
      index=1,
  )
  columns = get_products_sheet().keys()
  columns = list(columns)
  print("columns", columns)
  
  biggest_item = items.first()
  for item in items:
      if item.get_item_features().all().count() > biggest_item.get_item_features().all().count():
          biggest_item = item 
  for i in range(int(biggest_item.get_item_features().all().count())):
      columns.append("Название_Характеристики")
      columns.append("Измерение_Характеристики")
      columns.append("Значение_Характеристики")
  row_num = 1
  for col_num, column_title in enumerate(columns, 1):
      cell = worksheet1.cell(row=row_num, column=col_num)
      cell.value = column_title
  for item in items:
      discount = None
      if item.discount_type == 'v':
        discount = f'{item.discount}'
      elif item.discount_type == 'p':
        discount = f'{item.discount}'
        discount += '%'
      print(discount)
      if discount == "0.0":
        discount = None 
      print(discount)

      if item.unit and item.unit.name.lower() in units:
        unit = item.unit.name.lower()
      else:
        unit = 'шт.'

      from django.contrib.sites.models import Site 
      image_urls = []
      images = ItemImage.objects.filter(item=item)
      # print(ItemImage.objects.all())
      # print(images)
      # if images:
      #   print('!!!!!!!!')
      # if item.id == 1:
      #   print(item)
      #   print('item!')
      # print(item, images)
      for image in images:
        domain     = Site.objects.get_current().domain
        image_name = image.image.url
        image_url  = 'https://' + domain + image_name 
        image_urls.append(image_url)
      image_urls = ','.join(image_urls)
      
      category_id = None 
      if item.category:
        category_id = item.category.id 

      currency = 'UAH'      
      if item.currency:
        currency = item.currency.code

      manufacturer = None 
      if item.manufacturer:
        manufacturer = item.manufacturer.name
      
      availability = None
      if item.is_available:
        availability = '+'
      else:
        availability = '-'
      row_num += 1
      row = [
          # Код_товара
          item.id,
          # Название_позиции
          item.title,
          # Поисковые_запросы
          None,
          # Описание
          item.description,
          # Тип_товара
          'r',
          # Цена
          item.price,
          # Цена от
          None,
          # Ярлык
          None,
          # HTML_заголовок
          item.meta_title,
          # HTML_описание
          item.meta_descr,
          # HTML_ключевые_слова
          item.meta_key,
          # Валюта
          currency,
          # Скидка
          discount,
          # Cрок действия скидки от
          None,
          # Cрок действия скидки до
          None,
          # Единица_измерения
          unit,
          # Минимальный_объем_заказа
          None,
          # Оптовая_цена
          None,
          # Минимальный_заказ_опт
          None,
          # Ссылка_изображения
          image_urls,
          # Наличие
          availability,
          # Количество
          item.amount,
          # Производитель
          manufacturer,
          # Страна_производитель
          None,
          # Номер_группы
          None,
          # Адрес_подраздела
          None,
          # Возможность_поставки
          None,
          # Срок_поставки
          None,
          # Способ_упаковки
          None,
          # Личные_заметки
          None,
          # Продукт_на_сайте
          None,
          # Идентификатор_товара
          item.id,
          # Уникальный_идентификатор
          None,
          # Идентификатор_подраздела
          None,
          # Идентификатор_группы
          category_id,
          # Подарки
          None,
          # Сопутствующие
          None,
          # ID_группы_разновидностей
          None,
      ]
      features = []
      for feature in item.get_item_features().all():
          features.append(feature.name)
          features.append(feature.value)
      row.extend(features)
      for col_num, cell_value in enumerate(row, 1):
          cell = worksheet1.cell(row=row_num, column=col_num)
          if cell_value.__class__.__name__ == 'Feature':
            cell_value = cell_value.name 
          if cell_value.__class__.__name__ == 'FeatureValue':
            cell_value = cell_value.value 
          cell.value = cell_value
  return workbook


def create_worksheet_with_categories( workbook, categories):
  worksheet2 = workbook.create_sheet(
      title='Export Groups Sheet',
      index=2,
  )
  columns = get_categories_sheet().keys()
  row_num = 1
  for col_num, column_title in enumerate(columns, 1):
      cell = worksheet2.cell(row=row_num, column=col_num)
      cell.value = column_title
  for category in categories:
      row_num += 1
      parent_id = None 
      if category.parent:
        parent_id = category.parent.id
      row = [
          None, 
          # Номер_группы	
          category.title,
          # Название_группы	
          category.id, 
          # Идентификатор_группы	
          None,
          # Номер_родителя	
          parent_id,
          # Идентификатор_родителя	
          category.meta_title,
          # HTML_заголовок_группы	
          category.meta_descr,
          # HTML_описание_группы	
          category.meta_key,
          # HTML_ключевые_слова_группы
      ]
      for col_num, cell_value in enumerate(row, 1):
          cell = worksheet2.cell(row=row_num, column=col_num)
          cell.value = cell_value
  return workbook


def prom_export(request):
  response = HttpResponse(
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  )
  filename   = 'prom' # datetime.now().strftime('%Y-%m-%d')
  response['Content-Disposition'] = f"attachment; filename={filename}.xlsx"
  # exp        = ExportMixin()
  items      = Item.objects.all()
  categories = ItemCategory.objects.all() 
  workbook   = Workbook(); workbook.remove(workbook.active); 
  workbook   = create_worksheet_with_items(workbook, items)
  workbook   = create_worksheet_with_categories(workbook, categories)
  workbook.save(response)
  return response 


class GoogleProductsFeed(Rss201rev2Feed):
  content_type = 'application/xml; charset=utf-8'

  def rss_attributes(self):
      attrs = super().rss_attributes()
      attrs['xmlns:g'] = 'http://base.google.com/ns/1.0'
      attrs['xmlns:c'] = 'http://base.google.com/cns/1.0'
      return attrs

  def add_item_elements(self, handler, item):
    super().add_item_elements(handler, item)

    if item.get('id') is not None:
      handler.addQuickElement(u"g:id", item['id'])

    if item.get('mpn') is not None:
      handler.addQuickElement(u"g:mpn", item['mpn'])

    if item.get('condition') is not None:
      handler.addQuickElement(u"g:condition", item['condition'])

    if item.get('price') is not None:
      handler.addQuickElement(u"g:price", item['price'])

    if item.get('availability') is not None:
      handler.addQuickElement(u"g:availability", item['availability'])

    if item.get('brand') is not None:
      handler.addQuickElement(u"g:brand", item['brand'])

    if item.get('adult') is not None:
      handler.addQuickElement(u"g:adult", item['adult'])

    if item.get('product_type') is not None:
      handler.addQuickElement(u"g:product_type", item['product_type'])

    if item.get('multipack') is not None:
      handler.addQuickElement(u"g:multipack", item['multipack'])

    if item.get('image_links') is not None:
      key = u"g:image_link"
      for img in item['image_links']:
          handler.addQuickElement(key, img)
          key = u"g:additional_image_link"

    if item.get('product_details') is not None:
      for product_detail in item['product_details']:
        handler.startElement(u'g:product_detail', {})
        if product_detail.category:
          handler.addQuickElement(u"g:section_name", product_detail.category.name)
        handler.addQuickElement(u"g:attribute_name", product_detail.name.name)
        handler.addQuickElement(u"g:attribute_value", product_detail.value.value)
        handler.endElement(u'g:product_detail')

from django.contrib.sites.models import Site 
from .models import ItemImage 

class GoogleMerchant(Feed):
  title = "Faina Vishivanka"
  link = "/"
  # description = "Items from shop"
  feed_type = GoogleProductsFeed

  def items(self):
    # items = Item.objects.all() \
            # .filter(in_stock__availability=True) \ 
            # .filter(currency__isnull=False)
    # return items
    return Item.objects.all().filter(
      currency__isnull=False,
      in_stock__availability=True,
    )

  def item_title(self, item):
      return item.title

  def item_description(self, item):
      descr = strip_tags(item.description)
      return descr
  
  # def item_multipack(self, item):
  #   return item.multipack
  def item_guid(self, obj):
      return obj.id

  def item_extra_kwargs(self, item):
    image_links = []
    current_site = Site.objects.get_current().domain
    for image in ItemImage.objects.filter(item=item):
      image_links.append(f'https://{current_site}{image.image.url}')
    product_type = None 
    if item.category:
      product_type = item.category.tree_title.replace('->','>') 
    product_details = ItemFeature.objects.filter(item=item)
    # print(item.multipack, item)
    item_data = {
        "id":str(item.id),
        "mpn":str(item.id),
        "condition":"new",
        "price":f'{item.price} {item.currency.code.upper()}',
        "availability":"in_stock",
        "brand":getattr(item.brand, "title", None),
        "adult":"false",
        "product_type":product_type,
        "multipack": item.multipack,
        "image_links":image_links,
        "product_details":product_details,
    }
    return item_data




