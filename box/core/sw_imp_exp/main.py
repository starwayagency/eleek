from box.core.sw_imp_exp.main import * 

from django.http import HttpResponse
from django.core import serializers
from django.conf import settings 
from django.db import transaction
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import timedelta, datetime
from io import StringIO, BytesIO
from selenium import webdriver
from openpyxl import Workbook
from PIL import Image 

from bs4 import BeautifulSoup
from time import sleep
from random import choice, randint, uniform
from multiprocessing import Pool
import xlsxwriter
import random
import csv
import time
import csv
import ast 
import requests
import os 
import pandas as pd
import re 
import time 
import csv 
import re 

# from box.apps.sw_shop.sw_catalog.models import *
from box.core.sw_content.models import *

__all__ = [
  "Parser",
  "ExportMixin",
  "ImportMixin",
]


class Parser(object):
  categories_filename = 'categories.csv'
  products_filename   = 'products.csv'
  biggest_id          = 9000
  categories_links_json = 'categories_links.json'
  pages_links_json      = 'pages_links.json'
  products_links_json   = 'products_links.json'
  categories_links = [
    # {
    #   "number":"",
    #   "name":"",
    #   "id":"",
    #   "parent_id":"",
    #   "url":"",
    # },
  ]
  pages_links      = [
    # {
    #   "url":"",
    #   "category":"",
    # },
  ]
  products_links   = [
    # {
    #   "url":"",
    #   "category":"",
    # },
  ]
  # НЕИЗМЕНЯЕМОЕ(ОБЩЕЕ)

  def write_csv(self, data, filename):
    """Записывает информацию про товар в файл"""
    with open(filename, 'a') as file:
      writer = csv.DictWriter(file, fieldnames=list(data.keys()))
      writer.writerow(data)

  def write_json(self, data, filename, mode='a'):
    ''' Записывает в json-файл '''
    with open(filename, mode) as file:
      json.dump(data, file, indent=4, ensure_ascii=False)

  def get_proxies(self, ):
    """Возвращает список прокси-серверов"""
    html = requests.get('https://free-proxy-list.net/').text
    soup = BeautifulSoup(html, 'lxml')
    trs = soup.find('table', id='proxylisttable').find_all('tr')[10:50]
    proxies = []
    for tr in trs:
      tds = tr.find_all('td')
      ip = tds[0].text.strip()
      port = tds[1].text.strip()
      schema = 'https' if 'yes' in tds[6].text.strip() else 'http'
      proxy = {'schema': schema, 'address': ip + ':' + port}
      proxies.append(proxy)
    return choice(proxies)

  def get_html(self, url, proxies=False, headers=True):
    """Возвращает html-код страницы"""
    if proxies:
      p = get_proxies()
      proxies = { p['schema']: p['address']  }
    if headers:
      headers = headers
    else:
      headers = {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:66.0) Gecko/20100101 Firefox/66.0'}
    headers = {
      'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:66.0) Gecko/20100101 Firefox/66.0',
      'Host':"perkins-service.com.ua",
      'Referer':'https://perkins-service.com.ua/',
      'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
      # 'Connection':'keep-alive',
      'Accept-encoding':'en-US,en;q=0.5',

    }
    r = requests.get(url, proxies=proxies, headers=headers)
    return r.content if r.ok else print('ERROR OCCURED. STATUS CODE:',r.status_code)

    # Отримуємо проксі і юзерагент
    def fake_user(self, ):
        # proxies = open('proxies.txt').read().split('\n')
        useragents = open('user_agent.txt').read().split('\n')
        # proxie = choice(proxies)
        ua = choice(useragents)
        return ua

    # Задає параметри веб драйверу селеніума
    def proxy_driver(self, p, ua):
        co = webdriver.ChromeOptions()
        co.add_argument('--no-sandbox')
        co.add_argument('--disable-dev-shm-usage')
        co.add_argument(f"user-agent={ua}")
        co.add_argument("--headless")
        co.add_argument(f"--window-size={choice(('1024,768', '1280,960', '1280,1024', '1600,1200', '1920,1080'))}")
        pxy = p
        prox = Proxy()
        prox.proxy_type = ProxyType.MANUAL
        prox.http_proxy = pxy
        prox.socks_proxy = pxy
        prox.ssl_proxy = pxy

        capabilities = webdriver.DesiredCapabilities.CHROME
        # capabilities['resolution'] = '1920x1080'
        # prox.add_to_capabilities(capabilities)

        driver = webdriver.Chrome('/home/jurgeon/projects/jcb/materials/jcb/chromedriver', options = co, desired_capabilities = capabilities)

        return driver

    def authenticate(self, ):
        url = 'https://www.serwis-kop.pl/ru/authentication'

        try:
            driver = proxy_driver(
                choice(open('proxies.txt').read().split('\n')), 
                fake_user(),
            )
            print('driver created')
        except Exception as e:
            print(e)
            print('error driver create')
            quit()

        try:
            driver.get(url)
            # driver.save_screenshot('success.png')
            print('driver get url')
        except Exception as e:
            print(e)
            print('error get(url)')
            # driver.save_screenshot('error_geturl.png')
            driver.quit()

        try:
          element = WebDriverWait(driver, 13).until(EC.presence_of_element_located((By.ID, "SubmitLogin")))
        except Exception as e:
            print(e)
            print('bad html page')
            # driver.save_screenshot('bad_html.png')
            driver.quit()
        email = driver.find_element_by_id('email')
        email.send_keys('jurgeon018@gmail.com')

        passwd = driver.find_element_by_id('passwd')
        passwd.send_keys('yfpfhrj69018')

        submit = driver.find_element_by_id('SubmitLogin')
        print(submit)
        submit = driver.find_element_by_class_name('icon-lock')
        print(submit)
        # driver.save_screenshot('before_submit.png')
        submit.click()

        try:
            element = WebDriverWait(driver, 13).until(EC.presence_of_element_located((By.CLASS_NAME, "account")))
        except Exception as e:
            print(e)
            print('bad login')
            driver.save_screenshot('bad_login.png')
            driver.quit()

        # sleep(5)
        # driver.save_screenshot('login.png')
        html = driver.page_source
        return driver
        # driver.quit()

    def parse_products(self, link, driver):
      driver.get(link)
      soup         = BeautifulSoup(driver.page_source  , 'lxml')
      try:
        title        = soup.find('div', class_='primary_block').h1.text.strip()
        articul      = soup.find('p', id='product_reference').span.text.strip()
        manufacturer = soup.find_all('p', class_='product-detail-line')[-3].a.text.strip()
        price_netto  = soup.find_all('p', class_='product-detail-line')[-2].span.text.strip()
        price_brutto = soup.find_all('p', class_='product-detail-line')[-1].span.text.strip()
        description  = soup.find('section', class_='page-product-box').find('div', class_='content').text.strip()
        in_stock     = soup.find('div', class_='pb-center-column').find_all('p')[0].text.strip()
        imgs         = [li.a.get('href') for li in soup.find('ul', id="thumbs_list_frame").find_all('li')]
        categories   = []
        for meta in soup.find('div', class_='pshowbreadcrumb').find('div', class_='items').find_all('meta'):
          cat = re.sub(r'[^\w\s]+|[\d]+', r'', meta.get('content')).strip()
          if cat:
            categories.append(cat)
      except:
        title        = soup.find('div', class_='primary_block').h1.text
        articul      = soup.find('p', id='product_reference').span.text
        manufacturer = soup.find_all('p', class_='product-detail-line')[-1].a.text
        price_netto  = ''
        price_brutto = ''
        description  = soup.find('section', class_='page-product-box').find('div', class_='content').text.strip()
        in_stock     = soup.find('div', class_='pb-center-column').find_all('p')[0].text.strip()
        imgs         = [li.a.get('href') for li in soup.find('ul', id="thumbs_list_frame").find_all('li')]
        categories   = []
        for meta in soup.find('div', class_='pshowbreadcrumb').find('div', class_='items').find_all('meta'):
          cat = re.sub(r'[^\w\s]+|[\d]+', r'', meta.get('content')).strip()
          if cat:
            categories.append(cat)

      data = {
        'title':title,
        'articul':articul,
        'manufacturer':manufacturer,
        'price_netto':price_netto, 
        'price_brutto':price_brutto,
        'description':description,
        'in_stock':in_stock,
        'imgs':imgs,
        'categories':categories,
      }

      for key, value in data.items():
        print(f"{key}: {value}")
      print("________")
      
      write_csv(data, 'jcb_with_price.csv')

  def get_products_sheet(self, site='box', *args, **kwargs):
    if site == 'prom':
      fieldnames = {
        'Код_товара':'',
        'Название_позиции':'',
        'Поисковые_запросы':'',
        'Описание':'',
        'Тип_товара':'',
        'Цена':'',
        'Цена от':'',
        'Ярлык':'',
        'HTML_заголовок':'',
        'HTML_описание':'',
        'HTML_ключевые_слова':'',
        'Валюта':'',
        'Скидка':'',
        'Cрок действия скидки от':'',
        'Cрок действия скидки до':'',
        'Единица_измерения':'',
        'Минимальный_объем_заказа':'',
        'Оптовая_цена':'',
        'Минимальный_заказ_опт':'',
        'Ссылка_изображения':'',
        'Наличие':'',
        'Количество':'',
        'Производитель':'',
        'Страна_производитель':'',
        'Номер_группы':'',
        'Адрес_подраздела':'',
        'Возможность_поставки':'',
        'Срок_поставки':'',
        'Способ_упаковки':'',
        'Личные_заметки':'',
        'Продукт_на_сайте':'',
        'Идентификатор_товара':'',
        'Уникальный_идентификатор':'',
        'Идентификатор_подраздела':'',
        'Идентификатор_группы':'',
        'Подарки':'',
        'Сопутствующие':'',
        'ID_группы_разновидностей':'',
        'Название_Характеристики':'',
        'Измерение_Характеристики':'',
        'Значение_Характеристики':'',
        'Название_Характеристики':'',
        'Измерение_Характеристики':'',
        'Значение_Характеристики':'',
      }
    if site == 'box':
      fieldnames = {
        "Заголовок":"",
        "Описание":"",
        "Мета_Заголовок":"",
        "Мета_Описание":"",
        "Мета_Ключевые_Слова":"",
        "Артикул":"",
        "Категории":"",
        "Изображения":"",
        "Производитель":"",
        "Ссылка":"",
        "Валюта":"",
        "Старая_Цена":"",
        "Новая_Цена":"",
        "Наличие":"",
      }
    return fieldnames

  def get_categories_sheet(self, site='box', *args, **kwargs):
    if site == 'prom':    
      fieldnames = {
        'Номер_группы':'',
        'Название_группы':'',
        'Идентификатор_группы':'',
        'Номер_родителя':'',
        'Идентификатор_родителя':'',
        'HTML_заголовок_группы':'',
        'HTML_описание_группы':'',
        'HTML_ключевые_слова_группы':'',
      }
    if site == 'box':
      fieldnames = {
          "Заголовок":"",
          "Ссылка":"",
          "Ссылка_родителя":"",
          "Изображения":"",
      }

    return fieldnames

  def create_categories_json(self, url, *args, **kwargs):
    soup        = BeautifulSoup(self.get_html(url), 'lxml')
    categories  = soup.find('ul', class_='tree').find_all('a')
    cat    = soup.find('ul', class_='tree').find_all('a')[1]
    for cat in categories:

      try:
        raw         = cat.parent.parent.find_previous_sibling()
        parent_name = raw.text.strip()
        parent_url  = raw.get('href')
        with open('categories_links.json', 'r') as file:
          links     = json.load(file)['categories_links']
          for link in links:
            if link['name'].strip() == parent_name:
              parent_id = link['id']
      except Exception as e:
        print(e)
        parent_name = ''
        parent_url  = ''
        parent_id   = ''

      url  = cat.get('href')
      name = cat.text.strip()
      self.biggest_id += 1
      id = self.biggest_id
      category = {
        "id": id,
        "name":name,
        "url":url,
        "parent_name":parent_name,
        "parent_url":parent_url,
        "parent_id":parent_id,
      }
      # print('category: ', category)
      self.categories_links.append(category)
    self.write_json({'categories_links':self.categories_links}, 'categories_links.json', 'w')  

  def create_pages_json(self, url, start=1, stop=None, step=1):
    soup      = BeautifulSoup(self.get_html(url), 'lxml')
    category  = soup.find('div', class_="breadcrumb").text.split('>')[-1].strip()#.text.strip()
    try:
      last_page = soup.find('ul', class_='pagination').find_all('li')[-2].text
    except:
      last_page = 1
    page_pattern = '?p={i}'
    for i in range(1, int(last_page)+1):
      url  = url.split('?')[0]+page_pattern.format(i=i)
      with open('categories_links.json', 'r') as file:
        links = json.load(file)['categories_links']
        for link in links:
          if link['name'].strip() == category.strip():
            id = link['id']
      page = {
        "url":url,
        "category":category,
        "id":id,
      }
      print('page: ', page)
      self.page_links.append(page)
    self.write_json({'pages_links':self.page_links}, 'pages_links.json', 'w')  

  def create_products_links(self, link):
    url = link['url']
    category = link['category']
    id = link['id']
    soup = BeautifulSoup(self.get_html(url), 'lxml')

    # якшо сторінка - батьківська категорія, у якій є субкатегорії, 
    # то товари з неї не парсяться, бо вони потім будуть зустрічатись 
    # другий раз у дочірніх категоріях
    # if soup.find('div', id='subcategories') != None:
    if soup.find('li', class_='product-category') != None:
      return 
    # находим все блоки с товарами
    products = soup.find_all('li', class_='product')
    for product in products:
      link = product.a.get('href')
      product = {
        "url":link,
        "category":category,
        "id":id,
      }
      print('product: ', product)
      self.products_links.append(product)
    self.write_json({'products_links':self.products_links}, 'products_links.json', 'w')  
  
  # ! ПЕРЕОПРЕДЕЛИТЬ 

  def write_product_info(self, data):
    """Достает информацию про товар и записывает ее в файл"""
    url      = data.get('url', '')
    category = data.get('category', '')
    id       = data.get('id','')
    soup = BeautifulSoup(self.get_html(url), 'lxml')  
    name         = self.get_name(soup)
    desc         = self.get_description(soup)
    imgs         = self.get_imgs(soup)
    articule     = self.get_articule(soup, url)
    price        = self.get_price(soup)[0]
    currency     = self.get_price(soup)[1]
    availability = self.get_availability(soup)
    category     = self.get_category(soup, category)
    # category     = id
    data         = self.get_products_sheet()
    data         = self.get_features(soup, data)
    data['Код_товара']           = articule
    data['Название_позиции']     = name
    data['Описание']             = desc
    data['Тип_товара']           = 'r'
    data['Цена']                 = price
    data['Валюта']               = currency
    data['Единица_измерения']    = 'шт'
    data['Продукт_на_сайте']     = url
    data['Идентификатор_группы'] = category
    data['Идентификатор_товара'] = articule
    data['Ссылка_изображения']   = imgs
    data['Наличие']              = availability
    print("Ссылка: ",     url     )
    print("Изображения: ",imgs    )
    print("Название: ",   name    )
    print("Артикул: ",    articule)
    print("Цена: ",       price   )
    print("Валюта: ",     currency)
    print("Категория: ",  category)
    print("Описание: ",   desc    )
    self.write_csv(data=data, filename=self.products_filename)
  
  def get_name(self, soup):
    return name

  def get_category(self, soup, category):
    return category
    
  def get_description(self, soup):
    return desc

  def get_imgs(self, soup):
    return ', '.join(imgs)

  def get_articule(self, soup, *args):
    return articule

  def get_price(self, soup):
    return (price, currency)

  def get_availability(self, soup):
    return availability

  def get_features(self, soup, data):
    all_tds = []
    all_ths = []
    try:
      trs = soup.find('table', class_='woocommerce-product-attributes shop_attributes').find_all('tr')

      for tr in trs:
        ths = tr.find_all('th')
        tds = tr.find_all('td')
        for td in tds:
          all_tds.append(td.text.strip())
        for th in ths:
          all_ths.append(th.text.strip())
      names = [td for td in all_tds]
      values = [td for td in all_ths]
      # names = [td for td in all_tds if all_tds.index(td) % 2 == 0]
      # values = [td for td in all_tds if all_tds.index(td) % 2 != 0]


      features = dict(zip(names, values))
      params_dict = {}
      for key, value in features.items():
        params_dict.update({
          f'Характеристика({key})':key,
          f'Измерение({key})'     :'',
          f'Значение({key})'      :value
        })
      data.update(params_dict)
    except Exception as e:
      print('get_features()')
      print(e)
    return data

  # ! ПЕРЕОПРЕДЕЛИТЬ 

  def parse_categories(self, ):
    url = 'http://gvardeisky.com.ua/343-poncho-nakidka-dozhdevik-olive-germaniya.html'
    self.create_categories_json(url)

  def write_categories(self, ):
    with open('categories_links.json', 'r') as json_file:
      cats = json.load(json_file)['categories_links']
    with open('categories.csv', 'w') as csv_file:
      writer = csv.DictWriter(
        csv_file, 
        fieldnames=[k for k in self.get_categories_sheet().keys()]
      )
      writer.writeheader()
    with open('categories.csv', 'a') as csv_file:
      for cat in cats:
        name      = cat['name']
        id        = cat['id']
        parent_id = cat.get('parent_id', '')
        data      = self.get_categories_sheet()
        data['Название_группы']        = name
        data['Идентификатор_группы']   = id
        data['Идентификатор_родителя'] = parent_id
        self.write_csv(data, 'categories.csv')

  def parse_pages(self, ):
    with open('categories_links.json') as file:
      for category in json.load(file)['categories_links']:
        url = category['url'] + '?page=1'
        self.create_pages_json(url)

  def parse_products_links(self, ):
    parser = Parser()
    with open('pages_links.json','r') as file:
      for link in json.load(file)['pages_links']:
        parser.create_products_links(link)

  def write_products(self, turbo=False):
    parser = Parser()
    fieldnames = [k for k,v in parser.get_products_sheet().items()]
    with open('products.csv', 'w') as file:
      writer = csv.DictWriter(
        file, fieldnames=fieldnames
      )
      writer.writeheader()
    with open('products_links.json', 'r') as file:
      products = json.load(file)['products_links']
    if not turbo:
      for product in products:
        parser.write_product_info(product)
    else:
      with Pool(20) as pool:
        pool.map(
          parser.write_product_info, 
          [product for product in products]
        )

  def create_xlsx(self, ):
    workbook          = xlsxwriter.Workbook('result.xlsx') 
    products_sheet    = workbook.add_worksheet('Export Products Sheet')
    categories_sheet  = workbook.add_worksheet('Export Group Sheet')
    products_reader   = csv.reader(open('products.csv', 'r'), delimiter=',',quotechar='"')
    categories_reader = csv.reader(open('categories.csv', 'r'), delimiter=',',quotechar='"')
    
    row_count = 0
    for row in products_reader:
      for col in range(len(row)):
        # products_sheet.write(row_count,col,row[col])
        products_sheet.write_string(row_count,col,row[col])
      row_count +=1

    row_count = 0
    for row in categories_reader:
      for col in range(len(row)):
        # categories_sheet.write(row_count,col,row[col])
        categories_sheet.write_string(row_count,col,row[col])
      row_count +=1
    
    workbook.close()


class ExportMixin(Parser):

  def create_worksheet_with_items(self, workbook, items):
    worksheet1 = workbook.create_sheet(
        title='Export Products Sheet',
        index=1,
    )

    columns = self.get_products_sheet(site='box').keys()
    columns = list(columns)
    print("columns", columns)
    
    biggest_item = items.first()
    for item in items:
        if item.get_item_features().all().count() > biggest_item.get_item_features().all().count():
            biggest_item = item 
    for i in range(int(biggest_item.get_item_features().all().count())):
        columns.append("Название_Характеристики")
        columns.append("Значение_Характеристики")
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet1.cell(row=row_num, column=col_num)
        cell.value = column_title
    for item in items:
        row_num += 1
        row = [
            item.meta_title,
            item.meta_descr,
            item.meta_key,
            item.title,
            item.description,
            item.code,
            item.category.slug,
            item.slug,
            item.price,
            # item.old_price,
            item.price,
            ','.join([image.image.url for image in item.images.all()]),
        ]
        features = []
        for feature in item.get_item_features().all():
            features.append(feature.name)
            features.append(feature.value)
        row.extend(features)
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet1.cell(row=row_num, column=col_num)
            if cell_value.__class__.__name__ == 'Feature':
              cell_value = cell_value.name 
            if cell_value.__class__.__name__ == 'FeatureValue':
              cell_value = cell_value.value 
            cell.value = cell_value
    return workbook

  def create_worksheet_with_categories(self, workbook, categories):
    worksheet2 = workbook.create_sheet(
        title='Export Groups Sheet',
        index=2,
    )
    columns = self.get_categories_sheet(site='prom').keys()
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet2.cell(row=row_num, column=col_num)
        cell.value = column_title
    for category in categories:
        row_num += 1
        parent_id = None 
        if category.parent:
          parent_id = category.parent.id
        row = [
            category.title,
            # category.slug,
            # category.parent_slug,
            category.id,
            parent_id,

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = cell_value
    return workbook

  # terminal 

  def export_items_to_xlsx(self, request, queryset):
    items    = Item.objects.all()
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook = self.create_worksheet_with_items(workbook, items)
    workbook.save()
    with open(filename, 'w') as f:
      f.write(workbook)
    # TODO: нічо не зроблено

  def export_categories_to_csv(self, file_name=None, *args, **kwargs):
    if file_name == None:
      file_name = f"categories_{datetime.now().strftime('%Y-%m-%d')}.csv"
    data   = self.get_categories_sheet(site='box')
    for category in ItemCategory.objects.all():
      with open(file_name, 'a') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=[k for k in data.keys()])
        data['Заголовок']       = category.title
        data['Ссылка']          = category.slug
        data['Ссылка_родителя'] = category.parent_slug
        data['Изображения']     = category.image_path
        writer.writerow(data)
    return True 

  # ! terminal 

  # admin 

  def admin_export_items_to_xlsx(self, request, queryset):
      items = Item.objects.all()
      categories = ItemCategory.objects.all()
      response = HttpResponse(
          content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      )
      response['Content-Disposition'] = f"attachment; filename={datetime.now().strftime('%Y-%m-%d')}.xlsx"
      workbook = Workbook()
      workbook.remove(workbook.active)
      workbook = self.create_worksheet_with_items(workbook, items)
      workbook = self.create_worksheet_with_categories(workbook, categories)
      workbook.save(response)
      return response

  def admin_export_categories_to_csv(self, request, queryset):
    filename = f"categories_{datetime.now().strftime('%Y-%m-%d')}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachement; filename={filename}'
    data   = self.get_categories_sheet(site='box')
    writer = csv.writer(response)
    writer.writerow([k for k in data.keys()])
    for category in queryset:
      data['Заголовок']       = category.title
      data['Ссылка']          = category.slug
      data['Ссылка_родителя'] = category.parent_slug
      data['Изображения']     = category.image_path
      writer.writerow([value for value in data.values()])
    return response

  def admin_export_items_photoes(self, request, queryset):
    # https://thispointer.com/python-how-to-create-a-zip-archive-from-multiple-files-or-directory/
    import os 
    from zipfile import ZipFile, ZIP_DEFLATED
    from wsgiref.util import FileWrapper
    from django.conf import settings 

    images = []
    for item in queryset:
      images.extend(item.images.all() )
    print(len(images))

    with ZipFile('export.zip', 'w', ZIP_DEFLATED) as export_zip:
      for image in images:
        try:
          filename = settings.MEDIA_ROOT + image.image.url[6:]
          arcname = os.path.join('shop', 'item', image.item.slug, filename.split('/')[-1])
          export_zip.write(filename, arcname)
        except FileNotFoundError as e:
          print(e)
    response = HttpResponse(FileWrapper(open('export.zip', 'rb')), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=export.zip'
    return response 

  def admin_delete_items_photoes(self, request, queryset):
    for item in queryset:
      item.images.all().delete()

  def admin_delete_items_features(self, request, queryset):
    for item in queryset:
      item.get_item_features().all().delete()
  
  # ! admin 

  admin_export_items_to_xlsx.short_description     = "Експорт обраних товарів в ексель"
  admin_export_categories_to_csv.short_description = 'Експортувати обрані категорії '
  admin_export_items_photoes.short_description     = "Експорт у обраних товарів всіх фотографій"
  admin_delete_items_photoes.short_description     = "Видалити у обраних товарів всі фотографій"
  admin_delete_items_features.short_description    = 'Видалити у обраних товарів всі характеристики'

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  def export_items_to_xml(self, filename=None, *args, **kwargs):
    # https://github.com/vinitkumar/json2xml
    # https://github.com/quandyfactory/dicttoxml/tree/master/dist
    # !!!!
    # https://github.com/quandyfactory/dicttoxml
    # dicttoxml + rest_framework.serializers.ModelSerializer
    # https://stackoverflow.com/questions/36021526/converting-an-array-dict-to-xml-in-python
    # !!!!
    items = Item.objects.all()
    result = "<objects>"
    for item in items:
      result += form_item(item)
    result += "</objects>"
    # result  = serializers.serialize('xml', items)
    with open(filename, 'w') as f:
      f.write(result)
    return True

  def form_item(self, item):
    item = f"""
    <object>
      <field name="meta_title" descr="{item._meta.get_field('meta_title').verbose_name}">
        {item.meta_title}
      </field>
      <field name="meta_descr" descr="{item._meta.get_field('meta_descr').verbose_name}">
        {item.meta_descr}
      </field>
      <field name="meta_key" descr="{item._meta.get_field('meta_key').verbose_name}">
        {item.meta_key}
      </field>
      <field name="title" descr="{item._meta.get_field('title').verbose_name}">
        {item.title}
      </field>
      <field name="description" descr="{item._meta.get_field('description').verbose_name}">
        {item.description}
      </field>
      <field name="code" descr="{item._meta.get_field('code').verbose_name}">
        {item.code}
      </field>
      <field name="slug" descr="{item._meta.get_field('slug').verbose_name}">
        {item.slug}
      </field>
      <field name="image" descr="{item._meta.get_field('image').verbose_name}">
        {item.image}
      </field>
      <field name="old_price" descr="{item._meta.get_field('old_price').verbose_name}">
        {item.old_price}
      </field>
      <field name="price" descr="{item._meta.get_field('price').verbose_name}">
        {item.price}
      </field>
      <field name="currency" descr="{item._meta.get_field('currency').verbose_name}">
        {item.currency}
      </field>
      <field name="category" descr="{item._meta.get_field('category').verbose_name}">
        {item.category}
      </field>
      <field name="in_stock" descr="{item._meta.get_field('in_stock').verbose_name}">
        {item.in_stock}
      </field>
      <field name="is_new" descr="{item._meta.get_field('is_new').verbose_name}">
        {item.is_new}
      </field>
      <field name="is_active" descr="{item._meta.get_field('is_active').verbose_name}">
        {item.is_active}
      </field>
      <field name="created" descr="{item._meta.get_field('created').verbose_name}">
        {item.created}
      </field>
      <field name="updated" descr="{item._meta.get_field('updated').verbose_name}">
        {item.updated}
      </field>
      <field name="order" descr="{item._meta.get_field('order').verbose_name}">
        {item.order}
      </field>
    </object>
    """
    return item 


class ImportMixin(Parser):

  def read_items_from_xlsx(self, filename, *args, **kwargs):
    items      = pd.read_excel(filename, sheet_name="Товары")
    items      = items.to_csv()
    categories = pd.read_excel(filename, sheet_name="Категории")
    categories = categories.to_csv()
    create_categories(
      [dct for dct in map(dict, csv.DictReader(StringIO(categories)))],
      list(csv.reader(StringIO(categories))),
      *args,
      **kwargs,
    )
    create_items(
      [dct for dct in map(dict, csv.DictReader(StringIO(items)))], 
      list(csv.reader(StringIO(items))),
      *args,
      **kwargs,
    )
    return True


  def read_items_from_csv(self, filename, *args, **kwargs):
    create_items(
      [dct for dct in map(dict, csv.DictReader(open(filename)))], 
      list(csv.reader(open(filename))),
      *args,
      **kwargs,
    )
    return True 


  def read_items_categories_from_csv(self, filename, *args, **kwargs):
    self.create_items(
      [dct for dct in map(dict, csv.DictReader(open(filename)))], 
      list(csv.reader(open(filename))),
      *args,
      **kwargs,
    )
    return True 


  def read_categories_from_csv(self, filename, *args, **kwargs):
    self.create_categories(
      [dct for dct in map(dict, csv.DictReader(open(filename)))],
      list(csv.reader(open(filename))),
    )
    return True 


  def create_categories(self, categories, list_file, *args, **kwargs):
    # categories = categories[0:10]
    for category in categories:
      self.create_category(category)
    return True


  def create_category(self, category):
    title       = category["Заголовок"]
    descr       = category.get("Описание", title)
    meta_title  = category.get("Мета_Заголовок", title)
    meta_descr  = category.get("Мета_Описание", descr)
    meta_key    = category.get("Мета_Ключевые_Слова", title)
    code        = category.get("Код")
    slug        = category["Ссылка"]
    parent_slug = category["Ссылка_родителя"]
    image       = category.get('Изображение')
    from django.db.utils import IntegrityError
    try:
      new_category, _ = ItemCategory.objects.get_or_create(
        # slug__iexact = slug.lower().strip(),
        title = title.lower().strip(),
      )
      # new_category.title      = title
      new_category.code       = code
      new_category.meta_title = title
      new_category.meta_descr = meta_descr
      new_category.meta_key   = title
      new_category.description= descr
      parent = ItemCategory.objects.filter(
        slug = parent_slug, 
      ).first()
      new_category.parent = parent
      if code:
        new_category.code = code 
      if image:
        new_category.image = 'shop/category/' + image
      new_category.save()
      print(new_category)
    except IntegrityError as e:
      print(e)
      print(title)


  def parse_item_features(self, items, list_file, *args, **kwargs):
    items       = items
    headers_row = list_file[0]
    items_rows  = list_file[1:]
    feature_name_indexes  = [i for i, o in enumerate(headers_row) if "Название_Характеристики" in o]
    feature_value_indexes = [i for i, o in enumerate(headers_row) if "Значение_Характеристики" in o]
    # feature_code_indexes  = [i for i, o in enumerate(headers_row) if "feature_code" in o]
    items_features = []
    item_features_names  = []
    item_features_values = []
    # item_features_codes  = []
    features = []
    for item_row in items_rows:
      for i in feature_name_indexes:
        item_features_names.append(item_row[i])
      for i in feature_value_indexes:
        item_features_values.append(item_row[i])
      # for i in feature_code_indexes:
      #   item_features_codes.append(item_row[i])
      # TODO: допиляти так, шоб можна було в ексель-файл записувати 
      # крім назви характеристики і значення характеристики, ще й код 
      # характеристики і батьківську категорію
      # for i in range(len(item_features_names), *args, **kwargs):
      #   features.append({
      #     "feature_name": ,
      #     "":,
      #     "":,
      #     "":,
      #   })
      features = dict(zip(item_features_names, item_features_values))
      items_features.append(features) 

    for i in range(len(items_features), *args, **kwargs):
      items[i]["features"] = items_features[i]
    
    return items 


  def create_items(self, items, list_file, *args, **kwargs):
    items = self.parse_item_features(items, list_file, *args, **kwargs)
    # items = items[161+77+22+119+622:]
    # items = items[1464+1722:]
    # items = items[159:161]
    # try:
    for item in items:
      # return print('sdfsdf')
      try:
        new_item = self.create_item(item, *args, **kwargs)
        if new_item:
          self.print_item(item, new_item)
        print("items.index(item):", items.index(item))
      except Exception as e:
        print(e)
    # except Exception as e:
    #   print(e)


  def create_item(self, item, *args, **kwargs):
    title       = item["Заголовок"]
    description = item.get("Описание", title)
    meta_title  = item.get("Мета_Заголовок", title)
    meta_descr  = item.get("Мета_Описание", description)
    meta_key    = item.get("Мета_Ключевые_Слова", description)
    code        = item["Артикул"]
    existing_items = Item.objects.filter(code=code)
    if existing_items.exists() and not existing_items.first().title.lower().strip() == title.lower().strip():
        print(f'ITEM WITH CODE {code} ALREADY EXISTS')
        return 
    new_item, _ = Item.objects.get_or_create(
      # code=code,
      title=title,
    )
    # new_item.title       = title
    new_item.code       = code
    new_item.description = description
    new_item.meta_title  = meta_title
    new_item.meta_descr  = meta_descr
    new_item.meta_key    = meta_key
    new_item = self.handle_categories(item, new_item, *args, **kwargs)
    new_item = self.handle_slug(item, new_item, *args, **kwargs)
    new_item = self.handle_manufacturer(item, new_item, *args, **kwargs)
    new_item = self.handle_features(item, new_item, *args, **kwargs)
    new_item = self.handle_currency(item, new_item, *args, **kwargs)
    new_item = self.handle_price(item, new_item, *args, **kwargs)
    new_item = self.handle_in_stock(item, new_item, *args, **kwargs)
    # new_item = self.handle_images(item, new_item, *args, **kwargs)
    new_item.save()
    # return new_item 


  def print_item(self, item, new_item, *args, **kwargs):
    for field in new_item._meta.fields:
        print(field.name+':')
        print(getattr(new_item, field.name))
        print('----')
    print("__________________________")
    print("__________________________")
    for k, v in item.items():
        print(k+':')
        print(v)
        print('----')


  def handle_slug(self, item, new_item, *args, **kwargs):
    slug = item.get('Ссылка')
    if slug:
      if not Item.objects.filter(slug=slug).exists():
        new_item.slug = slug   
    return new_item


  def handle_manufacturer(self, item, new_item, *args, **kwargs):
    manufacturer = item.get('Производитель')
    if manufacturer:
      new_item.manufacturer, _ = ItemManufacturer.objects.get_or_create(name=manufacturer)
    return new_item


  def handle_images(self, item, new_item, *args, **kwargs):
    images = item.get("Изображения")
    if images:
      if not images[:2] == "['" and not images[-2:] == "']":
        images = images.split(",")
        for image in images:
          image = ItemImage.objects.create(
            image = f"shop/items/{image.strip()}",
            item  =  new_item, 
          )
        new_item.save()
        new_item.create_image_from_images()
      if images[:2] == "['" and images[-2:] == "']":
        images = ast.literal_eval(images)
        for image in images:
          ext = image.split('.')[-1]
          # new_image = ItemImage.objects.create(
          new_image, created = ItemImage.objects.get_or_create(
              item=new_item,
              image=f'shop/item/{new_item.slug}/{new_item.slug}created{images.index(image)}.{ext}'
          )
          print(new_image)
          # if created:
          if True:
            path = new_image.image.path
            save_path = '/'.join(path.split('/')[:-1])
            try:
                os.makedirs(save_path)
            except:
                pass
            image = Image.open(BytesIO(requests.get(image).content))
            print("image:  ")
            print(image)
            print(image)
            image.save(path)
            new_item.save()
            new_item.create_image_from_images()

    return new_item


  def handle_categories(self, item, new_item, *args, **kwargs):
    from django.db.utils import IntegrityError
    categories = item.get("Категории", "").lower().strip()
    if categories:
      if categories[0:2] == "['" and categories[-2:] == "']":
        categories   = ast.literal_eval(categories)
        if len(categories) == 3:
          if categories[0].lower().strip() == 'запчасти для jcb':
            for category_title in categories:
              try:
                category, _ = ItemCategory.objects.get_or_create(title=category_title.lower().strip())
                category_index = categories.index(category_title)
                if category_index != 0:
                  parent_category_title = categories[category_index-1]
                elif category_index == 0:
                  parent_category_title = "Запчастини та комплектуючі"
                parent_category, _ = ItemCategory.objects.get_or_create(title=parent_category_title.lower().strip())
                category.parent    = parent_category
                category.save()
                new_item.set_category([category,])
              except IntegrityError as e:
                print(e)
                print(categories)
      else:
        category = categories
        if category:
          parts = ItemCategory.objects.get(title="Запчастини та комплектуючі".lower().strip())

          parent = ItemCategory.objects.get(title='Запчастини PERKINS'.lower())
          # parent = ItemCategory.objects.get(code='parts_perkins')

          title = category.lower().strip()

          # print(title)
          # title = 'запчастини до двигуна perkins 400'#.lower().strip()
          # print(title)
          slug  = 'zapchasti-k-dvigatelju-perkins-400'

          perkinses_400_slug = ItemCategory.objects.filter(
            slug=slug
          )
          perkinses_400_title = ItemCategory.objects.filter(
            # title_ru=title,
            title_uk=title,
            # title=title,
            # title__iexact=title,
            # title__icontains=title,
          )
          perkinses_400_slug_title = perkinses_400_slug.first().title 
          # perkinses_400_title_title = perkinses_400_title.first().title 

          print()
          print("title: ", title)
          print("slug: ",  slug)
          print("perkinses_400_slug: ", perkinses_400_slug)
          print("perkinses_400_title: ", perkinses_400_title)
          print("perkinses_400_slug_title: ", perkinses_400_slug_title)
          # print("perkinses_400_title_title: ", perkinses_400_title_title)
          print()

          
          category, _ = ItemCategory.objects.get_or_create(
            # slug__iexact=category.lower().strip(),
            title__iexact=title,
            # title__icontains=title,
            # title_ru=title,
          )
          parent.parent = parts
          parent.save()
          category.parent = parent 
          category.save()
          new_item.set_category([category,])
    return new_item


  def handle_features(self, item, new_item, *args, **kwargs):
    for k, v in item["features"].items():
      new_feature, _ = ItemFeature.objects.get_or_create(
        item = new_item, 
        name =k,
        value=v,
      )
    return new_item


  def handle_currency(self, item, new_item, *args, **kwargs):
    currency    = item.get('Валюта','')
    if currency:
      new_item.currency, _ = Currency.objects.get_or_create(name=currency)
    return new_item 


  def handle_price(self, item, new_item, *args, **kwargs):
    old_price    = item.get("Старая_Цена")
    price    = item.get("Новая_Цена")
    price_netto  = item.get("price_netto")
    price_brutto = item.get("price_brutto")
    if old_price:
      new_item.old_price = old_price
    if price:
      new_item.price = price

    if '€' in price_netto:
      new_item.old_price   = float(price_netto.replace('€', '').strip().replace(',','.').replace(' ', ''))
      new_item.price   = float(price_brutto.replace('€', '').strip().replace(',','.').replace(' ', ''))
      new_item.currency, _ = Currency.objects.get_or_create(name='EUR')
    elif 'грн' in price_netto:
      price_netto          = price_netto.replace(' ', '').replace('грн.', '').strip().replace(',','.').replace(' ', '')
      price_brutto         = price_brutto.replace(' ', '').replace('грн.', '').strip().replace(',','.').replace(' ', '')
      price_netto          = price_netto.replace('\xa0', '').replace(' ', '')
      price_brutto         = price_brutto.replace('\xa0', '').replace(' ', '')
      new_item.old_price   = float(price_netto)
      new_item.price   = float(price_brutto)
      new_item.currency, _ = Currency.objects.get_or_create(name='UAH')
    elif 'Цену уточняйте' in price_netto:
      new_item.old_price   = None 
      new_item.price   = None 
      new_item.currency    = None 
    # print('1:', new_item.currency)
    return new_item 


  def handle_in_stock(self, item, new_item, *args, **kwargs):
    in_stock    = item.get("Наличие")
    if in_stock:
      st, _ = ItemStock.objects.get_or_create(text=in_stock)
      if in_stock.strip() in ['Товар не доступен', 'Нет в наличии']:
        st.availability = False 
      new_item.in_stock = st
    return new_item 


  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 

  # DANGER!! DANGER!! XML AREA!! 


  def import_items_from_xml_by_xmljson(self, filename, *args, **kwargs):
    from xmljson import badgerfish as bf
    from xml.etree.ElementTree import fromstring
    import json 
    items = '<p id="main">Hello<b>bold</b></p>'
    items = open(filename).read()
    items = bf.data(fromstring(items))
    items = json.dumps(items)
    items = json.loads(items)
    items = items['objects']['object']
    for item in items:
      fields = item['field']
      import pprint; pprint.pprint(fields)


  class XMLItem(object):
    meta_title  = None
    meta_descr  = None
    meta_key    = None
    title       = None
    description = None
    code        = None
    slug        = None
    image   = None
    old_price   = None
    price   = None
    currency    = None
    category    = None
    in_stock    = None
    is_new      = None
    is_active   = None
    created     = None
    updated     = None
    order       = None
    fields      = {}

    def __str__(self):
      return f'{self.fields}'


  def import_items_from_xml_by_xml_etree(self, filename, *args, **kwargs):
    import xml.etree.ElementTree as ET
    # *******
    # tree = ET.parse(filename)
    # xml_data = tree.getroot()W
    # xmlstr = ET.tostring(xml_data, encoding='utf8', method='xml')
    # data_dict = dict(xmltodict.parse(xmlstr))
    # print(data_dict)
    # with open('new_data_2.json', 'w+') as json_file:
    #     json.dump(data_dict, json_file, indent=4, sort_keys=True)
    # *******
    # https://stackoverflow.com/questions/45144645/parsing-xml-file-in-django


    tree = ET.parse(filename)
    root = tree.getroot()
    for att in root:
      fields = att.findall('field')#.text
      item = Item() 

      xml_item = XMLItem()
      for field in fields:
        name  = field.get('name')
        descr = field.get('descr')
        text  = field.text
        xml_item.fields.update({name:text})
        # setattr(xml_item, name, text)
      print(xml_item)

      title       = xml_item.fields['title']
      meta_title  = xml_item.fields['meta_title']
      meta_descr  = xml_item.fields['meta_descr']
      meta_key    = xml_item.fields['meta_key']
      title       = xml_item.fields['title']
      description = xml_item.fields['description']
      code        = xml_item.fields['code']
      slug        = xml_item.fields['slug']
      image   = xml_item.fields['image']
      old_price   = xml_item.fields['old_price']
      price   = xml_item.fields['price']
      currency    = xml_item.fields['currency']
      category    = xml_item.fields['category']
      in_stock    = xml_item.fields['in_stock']
      is_new      = xml_item.fields['is_new']
      is_active   = xml_item.fields['is_active']
      created     = xml_item.fields['created']
      updated     = xml_item.fields['updated']
      order       = xml_item.fields['order']
      print(xml_item.fields)
      return '123'



      # currency = Currency.objects.get()


      item.meta_title  = meta_title
      item.meta_descr  = meta_descr
      item.meta_key    = meta_key
      item.title       = title
      item.description = description
      item.code        = code
      item.slug        = slug
      item.image   = image
      item.old_price   = old_price
      item.price   = price
      item.currency    = currency
      item.category    = category
      item.in_stock    = in_stock
      item.is_new      = is_new
      item.is_active   = is_active
      item.created     = created
      item.updated     = updated
      item.order       = order
      # item.save()


  def import_items_from_xml_by_xmltodict(self, filename, *args, **kwargs):
    import xmltodict
    import json

    items = xmltodict.parse(open(filename).read())
    items = json.dumps(items)
    items = json.loads(items)
    items = items['objects']['object']
    for item in items:
      fields = item['field']
      for field in fields:
        item = Item()
        text = field.get('#text')
        name = field['@name']
        setattr(item, name, text)
        # item.save()
        print(item)
        import pprint; pprint.pprint(field)


  def import_items_from_xml(self, filename, *args, **kwargs):
    import_items_from_xml_by_xml_etree(filename)
    return True 

